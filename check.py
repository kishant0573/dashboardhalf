from flask import Flask, render_template, request, jsonify, redirect
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)
excel_file_path = 'Book1.xlsx'

@app.route('/')
def index():
    return render_template('check.html')

@app.route('/convert', methods=['POST'])
def convert():
    duration = request.form['duration']
    asr_solutions = request.form['asr_solutions']

    try:
        workbook = load_workbook(filename=excel_file_path)
        sheet = workbook.active

        last_row = sheet.max_row + 1

        sheet.cell(row=last_row, column=1).value = duration
        sheet.cell(row=last_row, column=2).value = asr_solutions

        workbook.save(filename=excel_file_path)

        return jsonify({'success': True})
    except Exception as e:
        return render_template('check.html', output=f"Error: {e}", success=False)


@app.route('/oktoopen', methods=['POST'])
def open_page():
    page_to_open = request.form['oktoopen']

    if page_to_open == 'page1':
        return redirect('/page1')
    else:
        return "Page not found"

@app.route('/page1')
def page1():
    try:
        excel_data = pd.read_excel(excel_file_path)
        json_data = excel_data.to_json(orient='records')
        return render_template('page1.html', json_data=json_data)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    app.run(debug=True)